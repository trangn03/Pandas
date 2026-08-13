import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. Load the Excel data 
file_path = r"M:\AS 9100 D\Approved Supplier List Certificates [INTERNAL]\F-11 APPROVED SUPPLIER LIST - 9.12.2025.xlsx"
df = pd.read_excel(file_path, header=3)

# 2. Clean the column headers
df.columns = df.columns.astype(str).str.replace('\n', ' ').str.strip()

# 3. Filter for Active suppliers only
active_suppliers = df[df['Select'] == 'Active'].copy()

# 4. Handle the mixed data and fill missing processes
active_suppliers = active_suppliers[active_suppliers['Date'] != 'Not Cert']
active_suppliers['Date'] = pd.to_datetime(active_suppliers['Date'], errors='coerce')
active_suppliers['Process'] = active_suppliers['Process'].fillna('N/A')

# 5. Define the expiration window (30 days from today)
today = pd.to_datetime('today')
warning_window = today + pd.Timedelta(days=30)

# 6. Find suppliers whose expiration date falls within the warning window
expiring_soon = active_suppliers[
    (active_suppliers['Date'] <= warning_window) & 
    (active_suppliers['Date'].notna())
].copy()

# Format the dates cleanly for our outputs
expiring_soon['Date'] = expiring_soon['Date'].dt.strftime('%m/%d/%y')

# --- LOCAL OUTPUTS START HERE ---

if not expiring_soon.empty:
    # 1. DISPLAY IN TERMINAL (Custom ASCII Table)
    print("\n" + "=" * 80)
    print(" ⚠️  ACTION REQUIRED: EXPIRING SUPPLIER CERTIFICATIONS  ⚠️ ".center(80))
    print("=" * 80)
    
    # Set up the column headers with specific spacing
    print(f" {'APPROVED SUPPLIER'.ljust(40)} | {'APPROVED PROCESS'.ljust(22)} | {'EXPIRES'} ")
    print("-" * 80)
    
    # Loop through and print each row perfectly aligned
    for index, row in expiring_soon.iterrows():
        # Truncate text if it's too long so it doesn't break the table structure
        supplier = str(row['Approved Supplier'])[:39].ljust(40)
        process = str(row['Process'])[:21].ljust(22)
        exp_date = row['Date']
        
        print(f" {supplier} | {process} | {exp_date} ")
        
    print("=" * 80)
    print(f" Total Expiring/Expired: {len(expiring_soon)}\n")
    
    # 2. SAVE TO ANOTHER FILE WITH FORMATTING
    output_filename = "Expiring_Suppliers_Check.xlsx"
    
    # Save the raw data using openpyxl engine
    expiring_soon.to_excel(output_filename, index=False, engine='openpyxl')
    
    # Load the workbook we just saved to apply styles
    wb = load_workbook(output_filename)
    ws = wb.active
    ws.title = "Expiring Certifications"

    # Define our quality reporting styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    header_font = Font(color="FFFFFF", bold=True) # White text
    border = Border(
        left=Side(style='thin', color='BFBFBF'), 
        right=Side(style='thin', color='BFBFBF'), 
        top=Side(style='thin', color='BFBFBF'), 
        bottom=Side(style='thin', color='BFBFBF')
    )
    date_font = Font(color="C00000", bold=True) # Red text for urgency

    # Apply styles to the Header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Apply borders to the data and autofit the column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        
        for cell in col:
            # Apply light borders to everything below the header
            if cell.row > 1:
                cell.border = border
                # Make the Expiration Date column bold red
                if ws.cell(row=1, column=cell.column).value == "Date":
                    cell.font = date_font
                    cell.alignment = Alignment(horizontal="center")
            
            # Find the longest text in each column
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                
        # Set the column width to fit the longest text perfectly
        ws.column_dimensions[col_letter].width = max_length + 4

    # Add dropdown filters and freeze the top row for easy scrolling
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Save the final formatted file
    wb.save(output_filename)
    print(f"[+] Formatted report successfully saved locally as: {output_filename}\n")

else:
    # Clean terminal output if nothing is expiring
    print("\n" + "=" * 60)
    print(" ✅ ALL CLEAR: NO EXPIRING CERTIFICATIONS ✅ ".center(60))
    print("=" * 60)
    print(" No active certifications are expiring within the next 30 days.\n")